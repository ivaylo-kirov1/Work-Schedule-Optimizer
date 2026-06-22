from __future__ import annotations

import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Employee, Schedule, ShiftAssignment, ShiftType

OFF_LABEL = "OFF"


_SHIFT_FILL_PALETTE = (
    "FFFFE0B2",  # light orange
    "FFB3E5FC",  # light blue
    "FFC8E6C9",  # light green
    "FFE1BEE7",  # light purple
    "FFFFF9C4",  # light yellow
    "FFFFCDD2",  # light red
)
_OFF_FILL_HEX = "FFEEEEEE"  # light gray

_EMPLOYEE_COLUMN_WIDTH = 24
_DAY_COLUMN_WIDTH = 11
_TITLE_ROW = 1
_SUBTITLE_ROW = 2
_HEADER_ROW = 4
_FIRST_DATA_ROW = _HEADER_ROW + 1


def _period_days(start: datetime.date, end: datetime.date) -> list[datetime.date]:
    span = (end - start).days
    return [start + datetime.timedelta(days=offset) for offset in range(span + 1)]


def _shift_fill_by_id(shift_types: list[ShiftType]) -> dict[int, PatternFill]:
    fills: dict[int, PatternFill] = {}
    for index, shift_type in enumerate(sorted(shift_types, key=lambda st: st.id)):
        color = _SHIFT_FILL_PALETTE[index % len(_SHIFT_FILL_PALETTE)]
        fills[shift_type.id] = PatternFill("solid", fgColor=color)
    return fills


def _assignment_map(
    schedule: Schedule, db: Session
) -> dict[tuple[int, datetime.date], int | None]:
    assignments = db.scalars(
        select(ShiftAssignment).where(ShiftAssignment.schedule_id == schedule.id)
    )
    return {
        (assignment.employee_id, assignment.date): assignment.shift_type_id
        for assignment in assignments
    }


def _write_titles(ws: Worksheet, schedule: Schedule) -> None:
    ws.cell(row=_TITLE_ROW, column=1, value=f"Work Schedule #{schedule.id}").font = (
        Font(bold=True, size=14)
    )
    subtitle = (
        f"Algorithm: {schedule.algorithm} · "
        f"Period: {schedule.start_date.isoformat()} to "
        f"{schedule.end_date.isoformat()}"
    )
    ws.cell(row=_SUBTITLE_ROW, column=1, value=subtitle)


def _write_header(ws: Worksheet, days: list[datetime.date]) -> None:
    bold = Font(bold=True)
    centered = Alignment(horizontal="center")
    employee_cell = ws.cell(row=_HEADER_ROW, column=1, value="Employee")
    employee_cell.font = bold
    for offset, day in enumerate(days):
        cell = ws.cell(
            row=_HEADER_ROW,
            column=2 + offset,
            value=f"{day.day} {day.strftime('%a')}",
        )
        cell.font = bold
        cell.alignment = centered


def _write_rows(
    ws: Worksheet,
    employees: list[Employee],
    days: list[datetime.date],
    cell_by_emp_date: dict[tuple[int, datetime.date], int | None],
    shift_index: dict[int, ShiftType],
    shift_fills: dict[int, PatternFill],
) -> None:
    off_fill = PatternFill("solid", fgColor=_OFF_FILL_HEX)
    centered = Alignment(horizontal="center")
    for row_offset, employee in enumerate(employees):
        row = _FIRST_DATA_ROW + row_offset
        name_cell = ws.cell(row=row, column=1, value=employee.name)
        name_cell.data_type = "s"  # force string: user-supplied, block formula injection
        for col_offset, day in enumerate(days):
            shift_type_id = cell_by_emp_date.get((employee.id, day))
            shift_type = shift_index.get(shift_type_id) if shift_type_id else None
            cell = ws.cell(row=row, column=2 + col_offset)
            if shift_type is None:
                cell.value = OFF_LABEL
                cell.fill = off_fill
            else:
                cell.value = shift_type.name
                cell.data_type = "s"  # force string: user-supplied, block formula injection
                cell.fill = shift_fills.get(shift_type.id, off_fill)
            cell.alignment = centered


def _apply_layout(ws: Worksheet, day_count: int) -> None:
    ws.column_dimensions["A"].width = _EMPLOYEE_COLUMN_WIDTH
    for offset in range(day_count):
        letter = get_column_letter(2 + offset)
        ws.column_dimensions[letter].width = _DAY_COLUMN_WIDTH
    ws.freeze_panes = f"B{_FIRST_DATA_ROW}"


def render_schedule_xlsx(schedule: Schedule, db: Session) -> bytes:
    employees = list(db.scalars(select(Employee).order_by(Employee.id)))
    shift_types = list(db.scalars(select(ShiftType)))
    shift_index = {st.id: st for st in shift_types}
    shift_fills = _shift_fill_by_id(shift_types)
    cell_by_emp_date = _assignment_map(schedule, db)
    days = _period_days(schedule.start_date, schedule.end_date)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Schedule"

    _write_titles(ws, schedule)
    _write_header(ws, days)
    _write_rows(ws, employees, days, cell_by_emp_date, shift_index, shift_fills)
    _apply_layout(ws, len(days))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
